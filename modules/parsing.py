"""
modules/parsing.py
Excel / CSV ingestion: classify sheet type, parse rows into list-of-dicts,
skip aggregate/totals rows.

CHANGED (sub-row match detail):
  _classify_subrow_cell() now returns (field_name, match_detail) so the
  transformation journey dialog can show exactly which regex or heuristic
  caused a sub-row cell to be assigned to "Address", "Cause of Loss", etc.
  _enrich_field() stores match_detail and subrow_inferred=True in the claim
  dict so dialogs.py can surface this in Step 2 of the field timeline.
"""

import csv
import os
import re

import openpyxl

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet classifier ──────────────────────────────────────────────────────────

def classify_sheet(rows) -> str:
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)

    if "line of business" in text:
        summary_co_signals = [
            "# claims", "num claims", "number of claims", "claim count",
            "loss ratio", "loss rate", "frequency", "severity",
        ]
        if any(sig in text for sig in summary_co_signals):
            return "SUMMARY"
        for row in rows[:20]:
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if non_empty and str(non_empty[0]).lower().strip() == "line of business" and len(non_empty) == 1:
                return "SUMMARY"

    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no", "file num",
        "file ref",
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date", "dol",
    ])
    has_fin = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ])
    if has_claim and (has_loss or has_fin):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


# ── Legacy-layout detector ────────────────────────────────────────────────────

def _is_legacy_print_layout(rows: list) -> bool:
    for row in rows:
        non_empty = [c for c in row if c is not None]
        if non_empty and all(str(c).strip() == "----------" for c in non_empty):
            return True

    for i in range(min(20, len(rows) - 1)):
        r1_vals = [str(c).strip() for c in rows[i] if c is not None]
        r2_vals = [str(c).strip() for c in rows[i + 1] if c is not None]
        if len(r2_vals) >= 5 and len(r1_vals) >= 2:
            r1_filled = sum(1 for c in rows[i] if c)
            r2_filled = sum(1 for c in rows[i + 1] if c)
            if r2_filled > r1_filled * 1.5 and r1_filled >= 2:
                combined = " ".join(r1_vals + r2_vals).lower()
                if ("file" in combined or "claim" in combined) and (
                    "paid" in combined or "incurred" in combined or "outstanding" in combined
                ):
                    return True
    return False


def _find_legacy_header_rows(rows: list) -> tuple[int, int] | None:
    for i in range(min(25, len(rows) - 1)):
        r1 = rows[i]
        r2 = rows[i + 1]
        r1_filled = sum(1 for c in r1 if c)
        r2_filled = sum(1 for c in r2 if c)
        if r2_filled < 4:
            continue
        combined = " ".join(
            str(c).lower() for c in list(r1) + list(r2) if c
        )
        if ("file" in combined or "claim" in combined or "assured" in combined) and (
            "paid" in combined or "outstanding" in combined or "incurred" in combined
        ):
            if r1_filled >= 2:
                return (i, i + 1)
            if r2_filled >= 5:
                return (i + 1, i + 1)
    return None


def _merge_two_header_rows(row1: list, row2: list) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for g, s in zip(row1, row2):
        g_s = str(g).strip() if g else ""
        s_s = str(s).strip() if s else ""
        if g_s and s_s and g_s.upper() != s_s.upper():
            name = f"{g_s} {s_s}"
        elif s_s:
            name = s_s
        elif g_s:
            name = g_s
        else:
            name = ""
        if name:
            seen[name] = seen.get(name, 0) + 1
            if seen[name] > 1:
                name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


# ── Sub-row / separator / subtotal detectors ─────────────────────────────────

def _is_separator_row(row_values: list) -> bool:
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if not non_empty:
        return False
    return all(str(c).strip() == "----------" for c in non_empty)


def _is_subtotal_row(row_values: list) -> bool:
    for c in row_values:
        if c is not None and str(c).strip():
            return bool(re.match(r"^total\b", str(c).strip(), re.IGNORECASE))
    return False


def _is_legacy_sub_row(row_values: list, num_cols: int) -> bool:
    if not row_values or row_values[0] is not None:
        return False
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if len(non_empty) == 0 or len(non_empty) > 3:
        return False
    has_addr_or_cause = (
        (len(row_values) > 1 and row_values[1] is not None) or
        (len(row_values) > 3 and row_values[3] is not None)
    )
    return has_addr_or_cause


# ── Smart sub-row cell classifier ─────────────────────────────────────────────

_ADDRESS_PAT = re.compile(
    r"""
    ^(
        \d+\s+\w.*            # "391 MAIN ST"
      | P\.?O\.?\s*BOX\s+\d   # "PO BOX 443"
      | \d+[-/]\d+\s+\w.*     # "12-14 ELM RD"
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

_ADDRESS_SUFFIX = re.compile(
    r"\b(st|street|ave|avenue|blvd|boulevard|rd|road|dr|drive|"
    r"ln|lane|ct|court|pl|place|way|cir|circle|hwy|highway|"
    r"pkwy|parkway|terr|ter|loop|trail|trl|run|box|suite|ste|"
    r"apt|unit|floor|fl)\b",
    re.IGNORECASE,
)

_CITY_STATE_ZIP_PAT = re.compile(
    r"""
    ^(
        [A-Za-z\s]{2,30},?\s+[A-Z]{2}\s+\d{5}(-\d{4})?   # City, ST 12345
      | [A-Za-z\s]{2,30},?\s+[A-Z]{2}$                     # City, ST
      | \d{5}(-\d{4})?$                                     # ZIP only
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

_CAUSE_OF_LOSS_PAT = re.compile(
    r"\b(fire|flood|wind|windstorm|hail|storm|tornado|hurricane|"
    r"tropical\s+storm|water\s+damage|water\s+intrusion|theft|"
    r"vandalism|slip|fall|trip|collision|accident|explosion|"
    r"lightning|freeze|ice|snow|earthquake|sinkhole|mold|"
    r"liability|negligence|assault|discrimination|wrongful|"
    r"product\s+liability|premises|auto|vehicle|medical|workers|"
    r"comp|injury|glass|burst\s+pipe|pipe\s+burst|roof|damage)\b",
    re.IGNORECASE,
)

_NAME_PAT = re.compile(
    r"^[A-Z][A-Za-z'-]+(\s+[A-Z][A-Za-z'-]+){1,4}$"
)


def _col_letter(col_index: int) -> str:
    result = ""
    n = col_index + 1
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


# ── CHANGED: returns (field_name, match_detail) ───────────────────────────────
def _classify_subrow_cell(value: str) -> tuple[str, str]:
    """
    Classify a single sub-row cell value into a semantic field name.

    Returns (field_name, match_detail) where match_detail is a human-readable
    description of which regex / heuristic caused the classification.
    This detail is forwarded all the way to the transformation journey dialog
    so reviewers can see exactly why a field was inferred (e.g. "regex: peril
    keyword matched — 'Tropical Storm'").

    field_name is one of:
      "Address"        – looks like a street address line
      "City State Zip" – looks like city/state/zip continuation
      "Cause of Loss"  – matches known peril / event vocabulary
      "Claimant Name"  – looks like a person/company name
      "Unknown"        – cannot be confidently classified
    """
    v = str(value).strip()
    if not v:
        return "Unknown", "empty value"

    if _ADDRESS_PAT.match(v):
        return "Address", "regex: street-number + name pattern (e.g. '391 MAIN ST', 'PO BOX …')"
    if _ADDRESS_SUFFIX.search(v) and re.search(r'\d', v):
        suffix_m = _ADDRESS_SUFFIX.search(v)
        suffix   = suffix_m.group(0) if suffix_m else ""
        return "Address", f"regex: street-suffix keyword '{suffix.upper()}' + digit present"
    if _CITY_STATE_ZIP_PAT.match(v):
        return "City State Zip", "regex: City ST 00000 / two-letter state / ZIP-only pattern"
    m = _CAUSE_OF_LOSS_PAT.search(v)
    if m:
        peril = m.group(0)
        return "Cause of Loss", f"regex: peril keyword matched — \"{peril}\""
    if _NAME_PAT.match(v) and not re.search(r'\d', v):
        return "Claimant Name", "regex: proper-noun name pattern (Title-cased words, no digits)"
    return "Unknown", "no pattern matched — stored verbatim"


# ── CHANGED: result tuple is now (value, col, match_detail) ──────────────────
def _infer_subrow_fields(raw_row: list) -> dict[str, tuple[str, int, str]]:
    """
    Given a raw sub-row (list of cell values), return a dict mapping
    inferred field names → (value, 1-based col index, match_detail).

    match_detail is the human-readable description of which regex/heuristic
    produced the classification — forwarded to _enrich_field() so the
    transformation journey dialog can show it.
    """
    result: dict[str, tuple[str, int, str]] = {}
    type_count: dict[str, int] = {}

    for c_idx, val in enumerate(raw_row):
        if val is None or not str(val).strip():
            continue

        val_s                  = str(val).strip()
        field_type, match_detail = _classify_subrow_cell(val_s)

        if field_type == "Unknown":
            field_type   = f"SubRow_{_col_letter(c_idx)}"
            match_detail = f"unclassified value in col {_col_letter(c_idx)} — stored verbatim"

        type_count[field_type] = type_count.get(field_type, 0) + 1
        if type_count[field_type] > 1:
            field_type = f"{field_type}_{type_count[field_type]}"

        result[field_type] = (val_s, c_idx + 1, match_detail)   # 1-based col

    return result


def _enrich_from_subrow(
    claim: dict,
    raw_row: list,
    r_idx: int,
) -> None:
    """
    Smart replacement for the hardcoded col-B/col-D sub-row enrichment.

    Infers field names from sub-row cell values using pattern matching and
    calls _enrich_field() for each one.  match_detail is forwarded so
    the transformation journey can show how each sub-row field was inferred.
    """
    inferred = _infer_subrow_fields(raw_row)
    for field_name, (value, excel_col, match_detail) in inferred.items():
        _enrich_field(
            claim, field_name, value,
            excel_row=r_idx, excel_col=excel_col,
            match_detail=match_detail,
        )


# ── Aggregate-row detection ───────────────────────────────────────────────────

_AGGREGATE_PATTERNS = re.compile(
    r"^(total|totals|grand\s*total|subtotal|aggregate|summary|sum|report\s*(date|total|summary)|"
    r"all\s+adjusters|ytd\s+total|period\s+total|fiscal\s+total|portfolio\s+total|"
    r"TOTALS_AGGREGATE|SUMMARY_FLIBBER|AGGREGATE_ZORP|SUMMARY_ZORP)",
    re.IGNORECASE,
)
_AGGREGATE_EXTRA = re.compile(
    r"(aggregate|zorp|flibber|summary|zoop|gorp|totals?_|_total|report_date|all_adjuster)",
    re.IGNORECASE,
)


def _is_aggregate_row(row_values: list) -> bool:
    non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return False
    first_val = non_empty[0]
    if _AGGREGATE_PATTERNS.match(first_val):
        return True
    if _AGGREGATE_EXTRA.search(first_val):
        return True
    first_tokens     = re.split(r"[_\s]+", first_val.lower())
    aggregate_tokens = {"total", "totals", "aggregate", "summary", "subtotal", "grand", "portfolio", "report"}
    if len(first_tokens) >= 2 and any(t in aggregate_tokens for t in first_tokens):
        return True
    for v in non_empty[:6]:
        if re.match(
            r"(total\s+claims|report\s+date|all\s+adjusters|open:\s*\d|pend:\s*\d|open:\d)",
            str(v), re.IGNORECASE,
        ):
            return True
    nums = [float(v) for v in row_values if isinstance(v, (int, float))]
    if nums and len(nums) >= 3 and all(n > 50_000 for n in nums):
        is_claim_id = (
            re.match(r"^[A-Z]{2,5}[-_][A-Z]{0,3}\d{3,}", first_val, re.IGNORECASE)
            or re.match(r"^\d{4,}$", first_val.strip())
        )
        if not is_claim_id:
            return True
    return False


# ── Sheet title / metadata extractor ─────────────────────────────────────────

_LABEL_ALIASES: dict[str, str] = {
    "prepared for":     "Reinsurer",
    "reinsurer":        "Reinsurer",
    "prepared by":      "TPA Name",
    "treaty":           "Treaty",
    "program":          "Treaty",
    "policy":           "Policy Number",
    "cedant":           "Cedant",
    "ceding company":   "Cedant",
    "insurer":          "Cedant",
    "valuation date":   "Valuation Date",
    "valuation":        "Valuation Date",
    "as of":            "Valuation Date",
    "report date":      "Report Date",
    "report generated": "Report Date",
    "effective date":   "Effective Date",
    "policy number":    "Policy Number",
    "policy no":        "Policy Number",
    "policy #":         "Policy Number",
    "insured":          "Insured Name",
    "named insured":    "Insured Name",
    "line of business": "Line of Business",
    "lob":              "Line of Business",
    "coverage":         "Coverage Type",
}


def _canonical_label(raw: str) -> str | None:
    key = raw.strip().rstrip(":").lower()
    return _LABEL_ALIASES.get(key)


def _try_inline_kv(cell_text: str) -> list[tuple[str, str]]:
    pairs = []
    segments = re.split(r'\s{3,}|\|', str(cell_text))
    for seg in segments:
        m = re.match(r'^([A-Za-z][^:]{0,40}):\s*(.+)$', seg.strip())
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
    return pairs


def extract_sheet_title_kvs(
    raw_rows: list,
    cell_rows: list,
    header_row_idx: int | None,
    sheet_name: str,
) -> dict:
    scan_limit = header_row_idx if header_row_idx is not None else min(15, len(raw_rows))
    found: dict = {}

    def _store(canonical: str, value: str, excel_row: int, excel_col: int):
        if canonical not in found and str(value).strip():
            found[canonical] = {
                "value":     str(value).strip(),
                "original":  str(value).strip(),
                "modified":  str(value).strip(),
                "source":    "title_kv",
                "excel_row": excel_row,
                "excel_col": excel_col,
            }

    for r_idx, row in enumerate(raw_rows[:scan_limit]):
        excel_row = r_idx + 1
        non_empty = [
            (c_idx, v) for c_idx, v in enumerate(row)
            if v is not None and str(v).strip()
        ]
        if not non_empty:
            continue

        if len(non_empty) == 1:
            c_idx, val = non_empty[0]
            val_s = str(val).strip()
            if re.match(r'^[\d$,()\-\.]+$', val_s):
                continue
            if ":" not in val_s:
                if r_idx == 0:
                    tpa_name = re.split(r'\s*[\u2014\u2013]\s*', val_s)[0].strip()
                    if ' - ' in tpa_name:
                        parts = tpa_name.split(' - ', 1)
                        if re.search(r'\b(report|run|detail|summary|schedule|listing)\b',
                                     parts[1], re.IGNORECASE):
                            tpa_name = parts[0].strip()
                    _store("TPA Name", tpa_name, excel_row, c_idx + 1)
                else:
                    lob_match = re.search(
                        r'(?:loss\s+run\s+report\s*[—\-–]+\s*'
                        r'|annual\s+loss\s+run\s*[—\-–]+\s*'
                        r'|program\s+year\s+\d{4}\s*[—\-–]?\s*)(.+)',
                        val_s, re.IGNORECASE,
                    )
                    if lob_match:
                        _store("Sheet Title", lob_match.group(1).strip(), excel_row, c_idx + 1)
                    else:
                        _store("Sheet Title", val_s, excel_row, c_idx + 1)
                continue

        for c_idx, val in non_empty:
            val_s = str(val).strip()
            if ":" in val_s and not re.match(r'^\d', val_s):
                for raw_label, raw_value in _try_inline_kv(val_s):
                    canonical = _canonical_label(raw_label)
                    if canonical:
                        _store(canonical, raw_value, excel_row, c_idx + 1)

        i = 0
        cells = non_empty
        while i < len(cells) - 1:
            c_label_idx, label_val = cells[i]
            c_value_idx, value_val = cells[i + 1]
            label_s = str(label_val).strip()
            value_s = str(value_val).strip()
            is_label = (
                label_s.endswith(":")
                or _canonical_label(label_s) is not None
            )
            if is_label and ":" in label_s and not label_s.endswith(":"):
                i += 1
                continue
            if is_label:
                canonical = (
                    _canonical_label(label_s.rstrip(":").strip())
                    or _canonical_label(label_s)
                )
                if canonical:
                    _store(canonical, value_s, excel_row, c_value_idx + 1)
                i += 2
            else:
                i += 1

    if "Sheet Name" not in found:
        found["Sheet Name"] = {
            "value":     sheet_name,
            "original":  sheet_name,
            "modified":  sheet_name,
            "source":    "sheet_tab",
            "excel_row": 0,
            "excel_col": 0,
        }

    return found


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_from_excel(
    file_path: str,
    sheet_name: str,
) -> tuple[list, str, dict]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN", {}
        claims, sheet_type = parse_rows(classify_sheet(rows), rows)
        return claims, sheet_type, {}
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN", {}

        sheet_type = classify_sheet(raw_rows)
        hri        = _find_header_row(raw_rows)
        title_kvs  = extract_sheet_title_kvs(raw_rows, cell_rows, hri, sheet_name)
        claims, sheet_type = parse_rows_with_cells(sheet_type, raw_rows, cell_rows)
        return claims, sheet_type, title_kvs


# ── Row parsers ───────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if (
            "claim" in rt or "employee name" in rt or "driver name" in rt
            or "claimant" in rt or "file" in rt
        ) and (
            "date" in rt or "incurred" in rt or "paid" in rt
            or "injury" in rt or "incident" in rt or "amount" in rt or "reserve" in rt
        ):
            return i
    for i, row in enumerate(rows[:5]):
        if sum(1 for c in row if c) >= 3:
            return i
    return None


def parse_rows_with_cells(sheet_type: str, rows: list, cell_rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
            r_idx = hri + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data: dict = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_with_cells(sheet_type, rows, cell_rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
        r_idx = hri + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total", "subtotal"] for c in raw_row if c):
            break
        if _is_aggregate_row(raw_row):
            continue
        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_with_cells(
    sheet_type: str, rows: list, cell_rows: list
) -> tuple[list, str]:
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(rows[top_hri])
            ]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(rows[i]) for i in range(len(rows))) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx_rel, (raw_row, cell_row) in enumerate(
        zip(rows[data_start:], cell_rows[data_start:])
    ):
        r_idx = data_start + 1 + r_idx_rel

        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            continue

        if _is_aggregate_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header = headers[c_idx_0]
            if not header:
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }

        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


# ── CHANGED: accepts match_detail kwarg ──────────────────────────────────────
def _enrich_field(
    claim: dict, field_name: str, value: str,
    excel_row: int, excel_col: int,
    match_detail: str = "",
) -> None:
    """Add or update a field in a claim dict if not already set.

    match_detail is stored in the claim dict so the transformation journey
    dialog can surface exactly how a sub-row field was inferred (e.g.
    'regex: peril keyword matched — "Tropical Storm"').
    subrow_inferred=True marks these fields as coming from a legacy sub-row
    rather than a header column.
    """
    if field_name not in claim or not claim[field_name].get("value"):
        claim[field_name] = {
            "value":           value,
            "modified":        value,
            "excel_row":       excel_row,
            "excel_col":       excel_col,
            "subrow_inferred": True,
            "match_detail":    match_detail or "sub-row pattern inference",
        }


# ── CSV / plain parse_rows (no cell objects) ──────────────────────────────────

def parse_rows(sheet_type: str, rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
            if not any(row):
                continue
            if _is_aggregate_row(list(row)):
                continue
            row_data: dict = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx - 1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_plain(sheet_type, rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        if _is_aggregate_row(list(row)):
            continue
        row_data: dict = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx - 1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_plain(sheet_type: str, rows: list) -> tuple[list, str]:
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[top_hri])]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(r) for r in rows) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            continue

        if _is_aggregate_row(list(raw_row)):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx, value in enumerate(raw_row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            if not header:
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


# ─────────────────────────────────────────────────────────────────────────────
# TXT / TRANSCRIPT PARSER
# ─────────────────────────────────────────────────────────────────────────────

def parse_txt_file(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a plain-text file (FNOL transcript, notes, etc.) into the
    structure expected by run_pdf_intelligence().
 
    Returns a dict with:
      - "full_text"   : entire decoded text  (primary — read by extract_full_text_from_parsed)
      - "pages"       : single-element list with page_num=1 and raw_text set
                        (kept for compatibility with _build_azure_di_index_from_parsed)
    """
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("latin-1", errors="replace")
 
    text = text.strip()
 
    return {
        "full_text": text,          # ← THIS is what extract_full_text_from_parsed reads
        "pages": [
            {
                "page_num":  1,
                "raw_text":  text,   # ← fallback if full_text key is not found
                "fields":    [],     # no ADI key-value pairs for TXT
                "page_width":  8.5,
                "page_height": 11.0,
            }
        ],
    }   


# ─────────────────────────────────────────────────────────────────────────────
# TXT / TRANSCRIPT PARSER   (appended to modules/parsing.py)
# Strategy: regex-first → LLM fallback if < MIN_REGEX_FIELDS found
# No hardcoded field names.
# ─────────────────────────────────────────────────────────────────────────────
 
import re as _re_txt
import os as _os_txt
import json as _json_txt
 
_TXT_MIN_REGEX_FIELDS = 3
 
# Generic colon-separated  "Label: value"
_RE_TXT_COLON = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{1,54}?)\s*:\s+(.+)$",
    _re_txt.MULTILINE,
)

# Generic equals-separated  "Label = value"
_RE_TXT_EQUALS = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{1,54}?)\s*=\s+(.+)$",
    _re_txt.MULTILINE,
)

# Table-style  "Label    Value"  (2+ spaces between)
_RE_TXT_TABLE = _re_txt.compile(
    r"^[ \t]*([A-Za-z][A-Za-z0-9 _./()'&-]{2,40})\s{2,}\t*([^\s].+)$",
    _re_txt.MULTILINE,
)
 
 
def _txt_is_plausible_label(label: str) -> bool:
    label = label.strip()
    if not label or len(label) > 55 or len(label.split()) > 7:
        return False
    if not label[0].isalpha():
        return False
    if _re_txt.search(r"\b(is|are|was|were|has|have|had|will|would|should|can|may)\b", label, _re_txt.I):
        return False
    return True
 
 
def _txt_regex_extract(text: str) -> list:
    seen: dict = {}
 
    def _nk(s: str) -> str:
        return _re_txt.sub(r"[\\s_\\-]+", "_", s.strip().lower())
 
    def _add(label: str, value: str, conf: float, raw: str) -> None:
        label = label.strip()
        value = value.strip()
        if not _txt_is_plausible_label(label) or not value or len(value) > 2000:
            return
        if value == value.upper() and len(value.split()) <= 4 and not _re_txt.search(r"\d", value):
            return
        nk = _nk(label)
        if nk not in seen:
            seen[nk] = {
                "field_name": label, "value": value,
                "confidence": conf, "source_text": raw.strip()[:200],
                "bounding_polygon": None, "source_page": 1,
                "page_width": None, "page_height": None,
                "excel_row": 1, "excel_col": None,
            }
 
    for m in _RE_TXT_COLON.finditer(text):
        _add(m.group(1), m.group(2), 0.90, m.group(0))
    for m in _RE_TXT_EQUALS.finditer(text):
        _add(m.group(1), m.group(2), 0.85, m.group(0))
    for m in _RE_TXT_TABLE.finditer(text):
        _add(m.group(1), m.group(2), 0.75, m.group(0))
 
    return list(seen.values())
 
 
def _txt_llm_extract(text: str) -> list:
    _system = (
        "You are a document field extractor. "
        "Extract every key-value pair from the document text. "
        "Return ONLY valid JSON: "
        '{"fields": [{"field_name": "...", "value": "...", "confidence": 0.0-1.0, "source_text": "..."}]}'
        " No markdown. No preamble. Use labels exactly as written. "
        "confidence: 0.95 if explicitly labelled, 0.75 if inferred."
    )
    try:
        from openai import AzureOpenAI
        client = AzureOpenAI(
            azure_endpoint=_os_txt.environ.get("OPENAI_DEPLOYMENT_ENDPOINT", ""),
            api_key=_os_txt.environ.get("OPENAI_API_KEY", ""),
            api_version=_os_txt.environ.get("OPENAI_API_VERSION", "2024-12-01-preview"),
        )
        model = _os_txt.environ.get("OPENAI_DEPLOYMENT_NAME", "gpt-4o-mini")
        truncated = text[:6000] + ("\n\n[... truncated ...]" if len(text) > 6000 else "")
        response = client.chat.completions.create(
            model=model, max_tokens=2500, temperature=0.0,
            messages=[
                {"role": "system", "content": _system},
                {"role": "user", "content": f"Extract fields from:\\n\\n{truncated}"},
            ],
        )
        raw = (response.choices[0].message.content or "").strip()
        raw = _re_txt.sub(r"^```(?:json)?\\s*", "", raw)
        raw = _re_txt.sub(r"\\s*```$", "", raw).strip()
        parsed = _json_txt.loads(raw)
        out = []
        for f in parsed.get("fields", []):
            if not f.get("field_name") or not f.get("value"):
                continue
            out.append({
                "field_name": str(f["field_name"]).strip(),
                "value": str(f["value"]).strip(),
                "confidence": float(f.get("confidence", 0.80)),
                "source_text": str(f.get("source_text", ""))[:200],
                "bounding_polygon": None, "source_page": 1,
                "page_width": None, "page_height": None,
                "excel_row": 1, "excel_col": None,
            })
        return out
    except Exception:
        return []
 
 
def _txt_merge(regex_fields: list, llm_fields: list) -> list:
    def _nk(s: str) -> str:
        return _re_txt.sub(r"[\\s_\\-]+", "_", (s or "").strip().lower())
    merged: dict = {}
    for f in llm_fields:
        merged[_nk(f["field_name"])] = f
    for f in regex_fields:
        nk = _nk(f["field_name"])
        if nk not in merged:
            merged[nk] = f
    return list(merged.values())
 
 
def parse_txt_file(file_bytes: bytes, filename: str) -> dict:
    """
    Parse a plain-text / transcript file.
 
    1. Decode bytes to text.
    2. Run generic regex extraction (colon, equals, tabular patterns).
    3. If regex finds < _TXT_MIN_REGEX_FIELDS: call LLM as fallback.
    4. Merge results (LLM priority, regex fills gaps).
 
    Returns dict compatible with run_pdf_intelligence():
      {"full_text", "pages", "doc_type", "doc_label", "source", ...}
    """
    text = ""
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue
    if not text:
        text = file_bytes.decode("utf-8", errors="replace")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
 
    regex_fields = _txt_regex_extract(text)
    _used_llm = False
 
    if len(regex_fields) < _TXT_MIN_REGEX_FIELDS:
        llm_fields = _txt_llm_extract(text)
        if llm_fields:
            fields = _txt_merge(regex_fields, llm_fields)
            _used_llm = True
        else:
            fields = regex_fields
    else:
        fields = regex_fields
 
    return {
        "full_text":           text,
        "source":              "txt",
        "doc_type":            "txt_document",
        "doc_label":           "Text Document",
        "_used_llm":           _used_llm,
        "_regex_field_count":  len(regex_fields),
        "pages": [{
            "page_num":   1,
            "page_label": "Page 1",
            "raw_text":   text,
            "fields":     fields,
        }],
    }    