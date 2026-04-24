"""
modules/excel_renderer.py
Renders an Excel sheet to a PIL Image and computes cell bounding boxes
for the eye-popup cell-highlight feature.

Also provides render_pdf_page_with_highlight() for PDF eye popup support.

FIXES:
  1. render_excel_sheet: removed duplicate draw.text call that caused every
     non-header cell to be drawn twice AND header row text to be skipped.
  2. get_cell_pixel_bbox: fixed zero-size bbox when target_col is the last
     column (col_starts[c] was clamped to col_starts[c-1]).
  3. render_excel_sheet: added guard so col_starts/row_starts are never empty
     even when openpyxl reports max_column/max_row as None or 0.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw

from modules.cell_format import _resolve_color, format_cell_value_with_fmt


def _col_px(ws, c: int, scale: float = 1.0) -> int:
    letter = get_column_letter(c)
    cd     = ws.column_dimensions.get(letter)
    w      = cd.width if (cd and cd.width and cd.width > 0) else 8.43
    return max(60, int(w * 10 * scale))


def _row_px(ws, r: int, scale: float = 1.0) -> int:
    rd = ws.row_dimensions.get(r)
    h  = rd.height if (rd and rd.height and rd.height > 0) else 15.0
    return max(14, int(h * 1.5 * scale))


def render_excel_sheet(excel_path: str, sheet_name: str, scale: float = 1.0):
    """Returns (PIL Image, col_starts, row_starts, merged_master)."""
    wb      = openpyxl.load_workbook(excel_path, data_only=True)
    ws      = wb[sheet_name]
    max_col = ws.max_column or 1
    max_row = ws.max_row    or 1

    # ── Build pixel start positions ──────────────────────────────────────────
    col_starts = [0]
    for c in range(1, max_col + 1):
        col_starts.append(col_starts[-1] + _col_px(ws, c, scale))
    row_starts = [0]
    for r in range(1, max_row + 1):
        row_starts.append(row_starts[-1] + _row_px(ws, r, scale))

    # Guard: ensure image is at least 1×1
    img_w = max(col_starts[-1], 1)
    img_h = max(row_starts[-1], 1)
    img  = Image.new("RGB", (img_w, img_h), "white")
    draw = ImageDraw.Draw(img, "RGBA")

    # ── Build merged-cell master map ─────────────────────────────────────────
    merged_master: dict = {}
    for mr in ws.merged_cells.ranges:
        mn_r, mn_c, mx_r, mx_c = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        for rr in range(mn_r, mx_r + 1):
            for cc in range(mn_c, mx_c + 1):
                merged_master[(rr, cc)] = (mn_r, mn_c, mx_r, mx_c)

    # ── Draw cells ───────────────────────────────────────────────────────────
    drawn_merges: set = set()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            merge_info = merged_master.get((r, c))
            if merge_info:
                mn_r, mn_c, mx_r, mx_c = merge_info
                if (mn_r, mn_c) in drawn_merges:
                    continue
                drawn_merges.add((mn_r, mn_c))
                x1 = col_starts[mn_c - 1]
                y1 = row_starts[mn_r - 1]
                x2 = col_starts[min(mx_c, len(col_starts) - 1)]
                y2 = row_starts[min(mx_r, len(row_starts) - 1)]
                cell = ws.cell(mn_r, mn_c)
            else:
                x1 = col_starts[c - 1]
                y1 = row_starts[r - 1]
                x2 = col_starts[c]          # col_starts has max_col+1 entries, so c is always valid
                y2 = row_starts[r]          # same for row_starts
                cell = ws.cell(r, c)

            # Background
            bg_hex = "FFFFFF"
            if cell.fill and cell.fill.fill_type == "solid":
                bg_hex = _resolve_color(cell.fill.fgColor, "FFFFFF")
            draw.rectangle([x1, y1, x2 - 1, y2 - 1],
                           fill=f"#{bg_hex}", outline="#CCCCCC", width=1)

            # Text — FIX: single draw.text call, chosen text based on row
            if cell.value is not None:
                txt_color = "#000000"
                if cell.font and cell.font.color:
                    fc = _resolve_color(cell.font.color, "000000")
                    if fc.upper() != bg_hex.upper():
                        txt_color = f"#{fc}"

                text  = format_cell_value_with_fmt(cell)
                bold  = bool(cell.font and cell.font.bold)

                if r == 1:
                    # Header row: never truncate
                    display_text = text
                else:
                    cell_w    = x2 - x1
                    ch_w      = 8 if bold else 7
                    max_chars = max(1, (cell_w - 8) // ch_w)
                    display_text = text[:max_chars - 1] + "…" if len(text) > max_chars else text

                # Single draw call — was duplicated before (caused double-paint
                # on non-header rows and missing paint on header row)
                draw.text((x1 + 4, y1 + 4), display_text, fill=txt_color)

    wb.close()
    return img, col_starts, row_starts, merged_master


def get_cell_pixel_bbox(
    col_starts: list, row_starts: list,
    target_row: int, target_col: int,
    merged_master: dict | None = None,
) -> tuple:
    """
    Return (x1, y1, x2, y2) pixel bounding box for the target cell.

    FIX: col_starts has (max_col + 1) entries so index `c` (1-based) is
    always a valid right-edge lookup.  The old code clamped to
    len(col_starts)-1 which aliased the last column's right edge to its
    left edge, producing a zero-width bbox.
    """
    n_cols = len(col_starts) - 1   # number of data columns available
    n_rows = len(row_starts) - 1

    c = max(1, min(target_col, n_cols))
    r = max(1, min(target_row, n_rows))

    if merged_master:
        info = merged_master.get((r, c))
        if info:
            mn_r, mn_c, mx_r, mx_c = info
            return (
                col_starts[max(0, mn_c - 1)],
                row_starts[max(0, mn_r - 1)],
                col_starts[min(mx_c, n_cols)],
                row_starts[min(mx_r, n_rows)],
            )

    return (
        col_starts[c - 1],
        row_starts[r - 1],
        col_starts[c],      # FIX: was col_starts[min(c, len-1)] which equalled col_starts[c-1] for last col
        row_starts[r],      # same fix for rows
    )


def crop_context(img, x1, y1, x2, y2, pad_x: int = 220, pad_y: int = 160):
    iw, ih   = img.size
    cx1, cy1 = max(0, x1 - pad_x), max(0, y1 - pad_y)
    cx2, cy2 = min(iw, x2 + pad_x), min(ih, y2 + pad_y)
    return img.crop((cx1, cy1, cx2, cy2)), x1 - cx1, y1 - cy1, x2 - cx1, y2 - cy1


# ─────────────────────────────────────────────────────────────────────────────
# PDF PAGE RENDERER WITH BOUNDING BOX HIGHLIGHT
# ─────────────────────────────────────────────────────────────────────────────

def render_pdf_page_with_highlight(
    pdf_path: str,
    page_number: int,
    bounding_polygon: list[tuple[float, float]] | None = None,
    page_width_inches: float = 8.5,
    page_height_inches: float = 11.0,
    dpi: int = 150,
) -> tuple:
    """
    Render a PDF page as a PIL Image with an optional bounding box highlight.

    Uses pymupdf (fitz) if available, falls back to pdf2image.

    Args:
        pdf_path:           Path to the PDF file.
        page_number:        1-based page number to render.
        bounding_polygon:   List of (x, y) coords in inches from Azure DI.
                            Should be 4 points forming a rectangle.
        page_width_inches:  Page width in inches (from Azure DI result).
        page_height_inches: Page height in inches (from Azure DI result).
        dpi:                Render resolution. 150 is fast and clear enough.

    Returns:
        (full_img, cropped_img) — both PIL Images.
        cropped_img is zoomed into the highlighted region with padding.
        Returns (None, None) if rendering fails.
    """
    img = None

    # ── Try pymupdf (fastest, no system deps) ────────────────────────────────
    try:
        import fitz  # pymupdf
        doc  = fitz.open(pdf_path)
        page = doc[page_number - 1]   # 0-based index
        mat  = fitz.Matrix(dpi / 72, dpi / 72)
        pix  = page.get_pixmap(matrix=mat, alpha=False)
        img  = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()
    except ImportError:
        pass
    except Exception:
        pass

    # ── Fallback: pdf2image (requires poppler) ────────────────────────────────
    if img is None:
        try:
            from pdf2image import convert_from_path
            pages = convert_from_path(
                pdf_path, dpi=dpi,
                first_page=page_number, last_page=page_number,
            )
            if pages:
                img = pages[0]
        except Exception:
            pass

    if img is None:
        return None, None

    img_w, img_h = img.size

    # If no bounding polygon, return full page without highlight
    if not bounding_polygon:
        return img, img

    # ── Convert Azure inch coords → pixel coords ──────────────────────────────
    scale_x = img_w / page_width_inches
    scale_y = img_h / page_height_inches
    px_poly = [(int(x * scale_x), int(y * scale_y)) for x, y in bounding_polygon]

    # ── Draw highlight on copy ────────────────────────────────────────────────
    highlighted = img.copy()
    draw = ImageDraw.Draw(highlighted, "RGBA")

    draw.polygon(px_poly, fill=(255, 230, 0, 90))
    for i in range(len(px_poly)):
        p1 = px_poly[i]
        p2 = px_poly[(i + 1) % len(px_poly)]
        draw.line([p1, p2], fill=(245, 158, 11, 255), width=3)
    for i in range(len(px_poly)):
        p1 = px_poly[i]
        p2 = px_poly[(i + 1) % len(px_poly)]
        draw.line([p1, p2], fill=(255, 255, 255, 160), width=1)

    xs = [p[0] for p in px_poly]
    ys = [p[1] for p in px_poly]
    x1, y1, x2, y2 = min(xs), min(ys), max(xs), max(ys)
    cropped, _, _, _, _ = crop_context(highlighted, x1, y1, x2, y2, pad_x=300, pad_y=200)

    return highlighted, cropped


def render_pdf_page_text_highlight(
    pdf_path: str,
    page_number: int,
    search_text: str,
    dpi: int = 150,
) -> tuple:
    """
    Render a PDF page and highlight the key AND its specific value together.
    search_text is expected to be in the format "KEY: VALUE" (source_text).
    Highlights key and value as a combined region, not every occurrence of value.
    Returns (full_img, cropped_img) as PIL Images, or (None, None) on failure.
    """
    try:
        import fitz
        from PIL import Image, ImageDraw

        doc  = fitz.open(pdf_path)
        page = doc[page_number - 1]

        key_text   = None
        value_text = None
        if ": " in search_text:
            parts      = search_text.split(": ", 1)
            key_text   = parts[0].strip()
            value_text = parts[1].strip()
        else:
            value_text = search_text.strip()

        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        doc.close()

        draw  = ImageDraw.Draw(img, "RGBA")
        scale = dpi / 72.0

        key_instances = []
        if key_text:
            doc2 = fitz.open(pdf_path)
            pg2  = doc2[page_number - 1]
            key_instances = pg2.search_for(key_text)

            val_instances = pg2.search_for(value_text) if value_text else []
            best_val = None
            if key_instances and val_instances:
                key_anchor_y = key_instances[0].y0
                best_val = min(val_instances, key=lambda r: abs(r.y0 - key_anchor_y))
            elif val_instances:
                best_val = val_instances[0]
            doc2.close()
        else:
            doc3 = fitz.open(pdf_path)
            pg3  = doc3[page_number - 1]
            val_instances = pg3.search_for(value_text)
            best_val = val_instances[0] if val_instances else None
            key_instances = []
            doc3.close()

        rects_to_highlight = []
        if key_instances:
            rects_to_highlight.append(key_instances[0])
        if best_val:
            rects_to_highlight.append(best_val)

        if rects_to_highlight:
            all_x1 = [r.x0 * scale for r in rects_to_highlight]
            all_y1 = [r.y0 * scale for r in rects_to_highlight]
            all_x2 = [r.x1 * scale for r in rects_to_highlight]
            all_y2 = [r.y1 * scale for r in rects_to_highlight]

            for rect in rects_to_highlight:
                x1 = rect.x0 * scale
                y1 = rect.y0 * scale
                x2 = rect.x1 * scale
                y2 = rect.y1 * scale
                draw.rectangle([x1, y1, x2, y2], fill=(255, 230, 0, 90))
                draw.rectangle([x1, y1, x2, y2], outline=(245, 158, 11, 255), width=3)
                draw.rectangle([x1 + 3, y1 + 3, x2 - 3, y2 - 3],
                               outline=(255, 255, 255, 160), width=1)

            cx1 = max(0, min(all_x1) - 300)
            cy1 = max(0, min(all_y1) - 80)
            cx2 = min(img.width,  max(all_x2) + 300)
            cy2 = min(img.height, max(all_y2) + 80)
            cropped = img.crop((cx1, cy1, cx2, cy2))
        else:
            cropped = img

        return img, cropped

    except Exception:
        return None, None